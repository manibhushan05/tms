import json
from datetime import datetime

from django.contrib import admin
from django.contrib.admin import site
from django.http import HttpResponse
from django.utils.encoding import smart_str
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from pygments import highlight
from pygments.formatters.html import HtmlFormatter
from pygments.lexers.data import JsonLexer
from rangefilter.filter import DateRangeFilter

from api.models import S3Upload
from broker.models import Broker
from restapi.models import BookingStatusesMapping
from sme.models import Sme
from supplier.models import Supplier
from team.models import ManualBooking, CreditDebitNoteReason, LrNumber
from team.payments.accounting import placed_order_accounting_summary, billed_customer_accounting_summary
from . import models


def outward_payment_status(booking):
    return booking.outward_booking.exists()


def inward_payment_status(booking):
    return booking.inward_booking.exists()


def invoice_status(booking):
    from team.models import Invoice
    if Invoice.objects.filter(bookings=booking).exists():
        return True
    else:
        return False


class OutWardPaymentAdmin(admin.ModelAdmin):
    list_display = ['id', 'bookings', 'adjusted_payment_id', 'lr_numbers', 'paid_to', 'actual_amount', 'payment_mode',
                    'status',
                    'payment_date', 'remarks', 'is_refund_amount', 'created_by', 'utr']
    list_display_links = ['bookings']
    list_filter = ('status', ('payment_date', DateRangeFilter), 'payment_mode', 'is_refund_amount', 'created_by')
    search_fields = ['id', 'paid_to', 'created_by__username', 'bank_account__account_number', 'fuel_card__card_number',
                     'lorry_number', 'payment_date', 'status', 'actual_amount', 'booking_id__booking_id',
                     'booking_id__lr_numbers__lr_number', 'remarks']
    autocomplete_fields = ['booking_id', 'bank_account', 'fuel_card', 'adjusted_outward_payments']
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def adjusted_payment_id(self, obj):
        return format_html('<br/>'.join(map(str, obj.adjusted_outward_payments.values_list('id', flat=True))))

    @staticmethod
    def bookings(obj):
        return "\n".join([booking.booking_id for booking in obj.booking_id.all()])

    @staticmethod
    def lr_numbers(obj):
        return format_html('\n'.join(['<br>'.join(booking.lr_numbers.values_list('lr_number', flat=True)) for booking in
                                      obj.booking_id.all()]), )

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status in ['paid', 'reconciled']:
            return False
        return True

    def delete_selected(self, request, obj):
        for value in obj:
            try:
                for booking in value.booking_id.all():
                    outward_amount = sum(
                        booking.outward_booking.values_list('actual_amount', flat=True))
                    booking.total_out_ward_amount = outward_amount
                    if outward_amount <= 0:
                        booking.outward_payment_status = 'no_payment_made'
                    booking.save()
                value.delete()
            except Exception as e:
                raise e

    delete_selected.short_description = 'Delete Selected Payments'


class InwardPaymentAdmin(admin.ModelAdmin):
    list_display_links = ['bookings']
    list_display = ['id', 'bookings', 'lr_numbers', 'received_from', 'actual_amount', 'tds', 'payment_mode',
                    'payment_date',
                    'remarks', 'created_on', 'created_by']
    search_fields = ['id', 'received_from', 'actual_amount', 'tds', 'payment_mode', 'remarks', 'invoice_number',
                     'booking_id__booking_id', 'booking_id__lr_numbers__lr_number', 'created_by__username']
    list_filter = (('payment_date', DateRangeFilter), 'payment_mode', 'created_by',)
    actions = ['delete_selected']
    autocomplete_fields = ['booking_id', 'created_by']
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    @staticmethod
    def bookings(obj):
        return "\n".join([booking.booking_id for booking in obj.booking_id.all()])

    @staticmethod
    def lr_numbers(obj):
        return format_html('\n'.join(['<br>'.join(booking.lr_numbers.values_list('lr_number', flat=True)) for booking in
                                      obj.booking_id.all()]), )

    def delete_selected(self, request, obj):
        for value in obj:
            try:
                for booking in value.booking_id.all():
                    inward_amount = sum(
                        booking.inward_booking.values_list('actual_amount', flat=True)) - value.actual_amount
                    booking.total_in_ward_amount = inward_amount
                    if inward_amount <= 0:
                        booking.inward_payment_status = 'no_payment'
                    elif inward_amount < booking.total_in_ward_amount:
                        booking.inward_payment_status = 'partial_received'
                    booking.tds_deducted_amount = sum(booking.inward_booking.values_list('tds', flat=True)) - value.tds
                    booking.save()
                value.delete()
            except Exception as e:
                raise e

    delete_selected.short_description = 'Delete Selected Payments'


class ManualBookingAdmin(admin.ModelAdmin):
    list_display_links = ['booking_id']
    list_filter = [('shipment_date', DateRangeFilter), 'invoice_status', 'pod_status', 'inward_payment_status',
                   'outward_payment_status', 'billing_type',
                   'booking_status', 'gst_liability', 'created_by']
    site.disable_action('delete_selected')
    actions = ['export_csv', 'export_xlsx']
    autocomplete_fields = ['source_office', 'destination_office', 'vehicle_category', 'from_city_fk', 'to_city_fk',
                           'company', 'supplier_vehicle', 'customer_to_be_billed_to', 'booking_supplier',
                           'supplier', 'owner_supplier', 'accounting_supplier', 'driver_supplier', 'created_by']
    readonly_fields = ['id', 'created_on', 'deleted_on', 'from_city', 'to_city','supplier','vehicle']

    list_display = (
        'id', 'booking_id', 'company_name', 'customer_to_be_billed', 'lr_numbers', 'lorry_number', 'shipment_date')
    search_fields = ['id', 'booking_id', 'company_code', 'lorry_number', 'shipment_date',
                     'driver_name', 'driver_phone', 'driver_dl_number', 'driver_dl_validity',
                     'truck_broker_owner_name', 'truck_broker_owner_phone', 'truck_owner_name', 'truck_owner_phone',
                     'company__name__profile__name', 'customer_to_be_billed_to__name__profile__name', 'booking_status',
                     'invoice_status', 'lr_numbers__lr_number']

    fieldsets = [
        ('Basic Details', {
            'fields': ['booking_id', 'billing_type', 'gst_liability', 'supplier_vehicle', 'vehicle_category',
                       'from_city_fk', 'to_city_fk', 'shipment_date', 'delivery_datetime', 'source_office',
                       'destination_office', 'company', 'customer_to_be_billed_to', 'company_code', 'created_by',
                       'booking_status_color']
        }),
        ('Rates', {
            'fields': ['loaded_weight', 'delivered_weight', 'charged_weight', 'party_rate',
                       'advance_amount_from_company', 'total_amount_to_company', 'total_in_ward_amount',
                       'tds_deducted_amount', 'supplier_charged_weight', 'supplier_rate', 'total_amount_to_owner',
                       'total_out_ward_amount', 'refund_amount']
        }),
        ('Status', {
            'fields': ['pod_status', 'pod_date', 'outward_payment_status', 'inward_payment_status', 'invoice_status',
                       'tds_certificate_status', 'booking_status', 'is_advance']
        }),
        ('Aaho Invoice', {
            'fields': ['to_be_billed_to', 'billing_address', 'billing_contact_number', 'billing_invoice_date',
                       'additional_charges_for_company', 'invoice_remarks_for_additional_charges',
                       'deductions_for_company', 'invoice_remarks_for_deduction_discount']
        }),
        ('Vendor Details', {
            'fields': ['driver_supplier', 'booking_supplier', 'accounting_supplier', 'owner_supplier']
        }),
        ('Vendor Charges', {
            'fields': ['loading_charge', 'unloading_charge', 'detention_charge', 'additional_charges_for_owner',
                       'note_for_additional_owner_charges']
        }),
        ('Vendor Deductions', {
            'fields': ['commission', 'lr_cost', 'deduction_for_advance', 'deduction_for_balance', 'other_deduction',
                       'remarks_about_deduction']
        }),
        ('Consignor Details', {
            'fields': ['consignor_name', 'consignor_address', 'consignor_city', 'consignor_city_fk', 'consignor_pin',
                       'consignor_phone',
                       'consignor_cst_tin', 'consignor_gstin']
        }),
        ('Consignee Details', {
            'fields': ['consignee_name', 'consignee_address', 'consignee_city', 'consignee_city_fk', 'consignee_pin',
                       'consignee_phone',
                       'consignee_cst_tin', 'consignee_gstin']
        }),
        ('Insurance', {
            'fields': ['is_insured', 'insurance_provider', 'insurance_policy_number', 'insured_amount',
                       'insurance_date', 'insurance_risk']
        }),
        ('Other Fields', {
            'fields': ['road_permit_number', 'party_invoice_number', 'party_invoice_date', 'party_invoice_amount',
                       'number_of_package', 'material', 'loading_points', 'is_print_payment_mode_instruction',
                       'unloading_points', 'remarks_advance_from_company', 'invoice_summary']
        })
    ]

    def has_add_permission(self, request):
        return False

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and (outward_payment_status(booking=obj) or inward_payment_status(booking=obj) or invoice_status(
                booking=obj) or obj.podfile_set.filter(verified=True, is_valid=True).exists()):
            return False
        return True

    def lr_numbers(self, obj):
        return format_html("<br>".join([lr.lr_number for lr in obj.lr_numbers.all()]))

    @staticmethod
    def company_name(obj):
        return '' if not obj.company else obj.company.get_company_name(obj.company_id)

    @staticmethod
    def customer_to_be_billed(obj):
        return '' if not obj.customer_to_be_billed_to else obj.customer_to_be_billed_to.get_company_name(
            obj.customer_to_be_billed_to_id)

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=Bookings {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow([smart_str(field.verbose_name) for field in ManualBooking._meta.fields])
        for obj in queryset:
            writer.writerow([smart_str(getattr(obj, field.name)) for field in ManualBooking._meta.fields])
        return response

    export_csv.short_description = u"Export CSV"

    def export_xlsx(modeladmin, request, queryset):
        import openpyxl
        try:
            from openpyxl.cell import get_column_letter
        except ImportError:
            from openpyxl.utils import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Bookings {}.xlsx'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "ManualBooking"

        row_num = 0

        columns = [(field.verbose_name, 20) for field in ManualBooking._meta.fields]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1

            row = [smart_str(getattr(obj, field.name)) for field in ManualBooking._meta.fields]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                # c.style.alignment.wrap_text = True

        wb.save(response)
        return response

    export_xlsx.short_description = u"Export XLSX"


class InvoiceAdmin(admin.ModelAdmin):
    list_display = ['id', 'invoice_number', 'booking_id', 'date', 'total_amount', 'company_name', 'customer_fk',
                    'created_by', 'payment_received', 'link', 'summary_required']
    list_display_links = ['invoice_number']
    list_filter = ['payment_received', 'summary_required', 'created_by']
    search_fields = ['id', 'invoice_number', 'company_name', 'bookings__booking_id']
    # actions = ['delete_selected']
    readonly_fields = ['id', 's3_upload', 'created_on', 'updated_on', 'deleted_on', 'deleted', 'changed_by',
                       'created_by']
    autocomplete_fields = ['bookings', 'customer_fk', 'city']

    def delete_selected(self, request, obj):
        for value in obj:
            try:
                for booking in value.bookings.all():
                    booking.invoice_status = 'no_invoice'
                    booking.save()
                    BookingStatusesMapping.objects.filter(
                        manual_booking=booking,
                        booking_status_chain__booking_status__status__in=[
                            'invoice_raised', 'invoice_confirmed',
                            'party_invoice_sent']).update(booking_stage='reverted')
            except ManualBooking.DoesNotExist:
                raise
            value.delete()

    delete_selected.short_description = 'Delete Selected Invoice'

    def link(self, obj):
        if isinstance(obj.s3_upload, S3Upload):
            return format_html('<a href="{}">download</a>'.format(obj.s3_upload.public_url()))
        else:
            return '-'

    link.short_description = 'Link'

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def booking_id(self, obj):
        return format_html("<br>".join([booking.booking_id for booking in obj.bookings.all()]))


class ToPayInvoiceAdmin(admin.ModelAdmin):
    list_filter = ['payment_received', 'created_by']
    list_display_links = ['invoice_number']
    list_display = ['id', 'invoice_number', 'booking_ids', 'company_name', 'date', 'total_payable_freight',
                    'payment_received', 'customer_fk', 'created_by']
    search_fields = ['id', 'invoice_number', 'company_name']
    actions = ['delete_selected']
    ordering = ('-id',)
    autocomplete_fields = ['bookings']
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def booking_ids(self, obj):
        return format_html("<br>".join([booking.booking_id for booking in obj.bookings.all()]))

    def delete_selected(self, request, obj):
        for value in obj:
            try:
                for booking in value.bookings.all():
                    booking.invoice_status = 'no_invoice'
                    booking.save()
            except ManualBooking.DoesNotExist:
                raise
            value.delete()

    delete_selected.short_description = 'Delete Selected Payments'


class LrNumberAdmin(admin.ModelAdmin):
    list_display = ['id', 'bookings', 'lr_number', 'datetime', 'source_office', 'destination_office', 'pod_status',
                    'created_by']
    search_fields = ['id', 'booking__booking_id', 'datetime', 'lr_number']
    readonly_fields = ['id', 'created_on', 'deleted_on', 'updated_on', 'deleted', 'created_by', 'changed_by']
    autocomplete_fields = ['booking', 'source_office', 'destination_office']

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    @staticmethod
    def bookings(obj):
        return obj.booking.booking_id if obj.booking else ''


class OutWardPaymentBillAdmin(admin.ModelAdmin):
    list_display_links = ['bookings']
    list_display = ['id', 'bookings', 'bill_number', 'bill_date', 'amount', 'vehicle_number', 'lr_number', 'from_city',
                    'to_city', 'loading_date', 'paid_to', 'outward_payment_ids', 'created_by']
    search_fields = ['id', 'booking__booking_id', 'bill_number', 'bill_date', 'amount', 'vehicle_number', 'lr_number',
                     'from_city', 'to_city', 'loading_date', 'paid_to']
    readonly_fields = ['id', 'outward_pmt', 'booking', 'changed_by', 'created_by', 'deleted_on', 'deleted',
                       'updated_on']

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    @staticmethod
    def outward_payment_ids(obj):
        return "\n".join([str(value.id) for value in obj.outward_pmt.all()])

    @staticmethod
    def bookings(obj):
        return obj.booking.booking_id


class DeletedDataAdmin(admin.ModelAdmin):
    list_display = ['id', 'model', 'data_prettified']
    search_fields = ['id', 'model', 'data']
    list_filter = ('model',)
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on')

    def data_prettified(self, instance):
        """Function to display pretty version of our data"""

        # Convert the data to sorted, indented JSON
        response = json.dumps(instance.data, sort_keys=True, indent=2)

        # Truncate the data. Alter as needed
        response = response

        # Get the Pygments formatter
        formatter = HtmlFormatter(style='colorful')

        # Highlight the data
        response = highlight(response, JsonLexer(), formatter)

        # Get the stylesheet
        style = "<style>" + formatter.get_style_defs() + "</style><br>"

        # Safe the output
        return mark_safe(style + response)

    data_prettified.short_description = 'data prettified'

    def delete_model(modeladmin, request, queryset):
        for obj in queryset:
            obj.delete()


class PendingInwardPaymentEntryAdmin(admin.ModelAdmin):
    list_display = [
        'id', 'customer_name', 'booking_ids', 'payment_ids', 'payment_mode', 'amount', 'tds', 'payment_date',
        'adjusted_flag',
        'uploaded_datetime',
        'uploaded_by', 'adjusted_datetime', 'adjusted_by',
    ]
    search_fields = [
        'id','customer__name__profile__name', 'payment_mode', 'amount', 'payment_date', 'adjusted_flag',
        'uploaded_datetime', 'uploaded_by__profile__name', 'adjusted_datetime', 'adjusted_by__profile__name',
        'bookings__booking_id', 'trn'
    ]
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')
    list_filter = ['adjusted_flag', 'payment_mode', 'adjusted_by']
    autocomplete_fields = ('bookings', 'inward_payment', 'customer', 'uploaded_by', 'adjusted_by')

    def booking_ids(self, obj):
        return format_html("<br>".join([booking.booking_id for booking in obj.bookings.all()]))

    def payment_ids(self, obj):
        return format_html("<br>".join([str(payment.id) for payment in obj.inward_payment.all()]))


class InvoiceSummaryAdmin(admin.ModelAdmin):
    list_display = ['id', 'ref_number', 'booking_ids', 'link', 'created_by', 'created_on', 'deleted'
                    ]
    search_fields = ['id', 's3_upload__filename']
    actions = ['delete_selected']
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 's3_upload')

    autocomplete_fields = ['created_by']

    def booking_ids(self, obj):
        return format_html("<br>".join([booking.booking_id for booking in obj.booking.all()]))

    def link(self, obj):
        if isinstance(obj.s3_upload, S3Upload):
            return format_html('<a href="{}">download</a>'.format(obj.s3_upload.public_url()))
        else:
            return '-'

    link.short_description = 'Link'


class CreditDebitNoteReasonAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'created_on')
    search_fields = ('id', 'name')
    list_filter = ('created_on',)
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True


class CreditNoteCustomerAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'credit_note_number', 'booking_id', 'customer', 'credit_amount', 'adjusted_amount', 'reason', 'status',
        'created_on', 'approved_on', 'adjusted_on')
    autocomplete_fields = ('bookings', 'customer', 'reason', 'approved_by', 'adjusted_by', 'invoice')
    search_fields = (
        'id', 'bookings__booking_id', 'customer__name__profile__name', 'credit_note_number',
        'bookings__lr_numbers__lr_number', 'reason__name', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    list_filter = (
        'status', 'deleted', 'reason', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 'credit_note_number')
    actions = ['export_csv']

    def booking_id(self, instance):
        return format_html('<br/>'.join([booking.booking_id for booking in instance.bookings.all()]))

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=CNC {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(
            ['id', 'credit_note_number', 'booking_id', 'customer', 'credit_amount', 'adjusted_amount', 'reason',
             'status', 'created_on', 'approved_on', 'adjusted_on', 'remarks'])
        for instance in queryset:
            writer.writerow([
                instance.id,
                instance.credit_note_number,
                '\n'.join([booking.booking_id for booking in instance.bookings.all()]),
                instance.customer.get_name() if isinstance(instance.customer, Sme) else None,
                instance.credit_amount,
                instance.adjusted_amount,
                instance.reason.name if isinstance(instance.reason, CreditDebitNoteReason) else None,
                instance.get_status_display(),
                instance.created_on.date() if instance.created_on else None,
                instance.approved_on.date() if instance.approved_on else None,
                instance.adjusted_on.date() if instance.adjusted_on else None,
                instance.remarks
            ])
        return response

    export_csv.short_description = u"Export CSV"


class DebitNoteCustomerAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'debit_note_number', 'booking_id', 'customer', 'debit_amount', 'adjusted_amount', 'reason', 'status',
        'created_on', 'approved_on', 'adjusted_on')
    autocomplete_fields = ('bookings', 'customer', 'reason', 'approved_by', 'adjusted_by', 'invoice')
    search_fields = (
        'id', 'bookings__booking_id', 'customer__name__profile__name', 'debit_note_number',
        'bookings__lr_numbers__lr_number', 'reason__name', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    list_filter = (
        'status', 'deleted', 'reason', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 'debit_note_number')
    actions = ['export_csv']

    def booking_id(self, instance):
        return format_html('<br/>'.join([booking.booking_id for booking in instance.bookings.all()]))

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=DNC {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(
            ['id', 'debit_note_number', 'booking_id', 'customer', 'debit_amount', 'adjusted_amount', 'reason',
             'status', 'created_on', 'approved_on', 'adjusted_on', 'remarks'])
        for instance in queryset:
            writer.writerow([
                instance.id,
                instance.debit_note_number,
                '\n'.join([booking.booking_id for booking in instance.bookings.all()]),
                instance.customer.get_name() if isinstance(instance.customer, Sme) else None,
                instance.debit_amount,
                instance.adjusted_amount,
                instance.reason.name if isinstance(instance.reason, CreditDebitNoteReason) else None,
                instance.get_status_display(),
                instance.created_on.date() if instance.created_on else None,
                instance.approved_on.date() if instance.approved_on else None,
                instance.adjusted_on.date() if instance.adjusted_on else None,
                instance.remarks
            ])
        return response

    export_csv.short_description = u"Export CSV"


class CreditNoteSupplierAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'credit_note_number', 'booking_id', 'accounting_supplier', 'credit_amount', 'adjusted_amount', 'reason', 'status',
        'created_on', 'approved_on', 'adjusted_on')
    autocomplete_fields = ('bookings', 'accounting_supplier', 'reason', 'approved_by','rejected_by', 'adjusted_by', 'invoice')
    search_fields = (
        'id', 'bookings__booking_id', 'broker__name__profile__name', 'credit_note_number',
        'bookings__lr_numbers__lr_number', 'reason__name', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    list_filter = (
        'status', 'deleted', 'reason', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 'credit_note_number','broker')
    actions = ['export_csv']

    def booking_id(self, instance):
        return format_html('<br/>'.join([booking.booking_id for booking in instance.bookings.all()]))

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=CNS {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(
            ['id', 'credit_note_number', 'booking_id', 'supplier', 'credit_amount', 'adjusted_amount', 'reason',
             'status', 'created_on', 'approved_on', 'adjusted_on', 'remarks'])
        for instance in queryset:
            writer.writerow([
                instance.id,
                instance.credit_note_number,
                '\n'.join([booking.booking_id for booking in instance.bookings.all()]),
                instance.accounting_supplier.name if isinstance(instance.accounting_supplier, Supplier) else None,
                instance.credit_amount,
                instance.adjusted_amount,
                instance.reason.name if isinstance(instance.reason, CreditDebitNoteReason) else None,
                instance.get_status_display(),
                instance.created_on.date() if instance.created_on else None,
                instance.approved_on.date() if instance.approved_on else None,
                instance.adjusted_on.date() if instance.adjusted_on else None,
                instance.remarks
            ])
        return response

    export_csv.short_description = u"Export CSV"


class CreditNoteCustomerDirectAdvanceAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'credit_note_number', 'booking_id', 'customer', 'accounting_supplier', 'credit_amount', 'adjusted_amount', 'reason',
        'status',
        'created_on', 'approved_on', 'adjusted_on')
    autocomplete_fields = ('bookings', 'accounting_supplier', 'reason', 'approved_by','rejected_by', 'adjusted_by', 'invoice', 'customer')
    search_fields = (
        'id', 'bookings__booking_id', 'customer__name__profile__name', 'customer__company_code',
        'broker__name__profile__name', 'credit_note_number',
        'bookings__lr_numbers__lr_number', 'reason__name', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    list_filter = (
        'status', 'deleted', 'reason', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 'credit_note_number','broker')
    actions = ['export_csv']

    def booking_id(self, instance):
        return format_html('<br/>'.join([booking.booking_id for booking in instance.bookings.all()]))

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=CNSA {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(
            ['id', 'credit_note_number', 'booking_id', 'supplier', 'customer', 'credit_amount', 'adjusted_amount',
             'reason', 'status', 'created_on', 'approved_on', 'adjusted_on', 'remarks'])
        for instance in queryset:
            writer.writerow([
                instance.id,
                instance.credit_note_number,
                '\n'.join([booking.booking_id for booking in instance.bookings.all()]),
                instance.accounting_supplier.name if isinstance(instance.accounting_supplier, Supplier) else None,
                instance.customer.get_name() if isinstance(instance.customer, Sme) else None,
                instance.credit_amount,
                instance.adjusted_amount,
                instance.reason.name if isinstance(instance.reason, CreditDebitNoteReason) else None,
                instance.get_status_display(),
                instance.created_on.date() if instance.created_on else None,
                instance.approved_on.date() if instance.approved_on else None,
                instance.adjusted_on.date() if instance.adjusted_on else None,
                instance.remarks
            ])
        return response

    export_csv.short_description = u"Export CSV"


class DebitNoteSupplierDirectAdvanceAdvanceAdmin(admin.ModelAdmin):
    list_display = ('id',)
    autocomplete_fields = (
        'bookings', 'supplier', 'customer', 'reason', 'approved_by','rejected_by', 'adjusted_by', 'invoice')
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by','broker')

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True


class DebitNoteSupplierAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'debit_note_number', 'booking_id', 'accounting_supplier', 'debit_amount', 'adjusted_amount', 'reason', 'status',
        'created_on', 'approved_on', 'adjusted_on')
    autocomplete_fields = ('bookings', 'accounting_supplier', 'reason', 'approved_by','rejected_by', 'adjusted_by', 'invoice')
    search_fields = (
        'id', 'bookings__booking_id', 'broker__name__profile__name', 'debit_note_number',
        'bookings__lr_numbers__lr_number', 'reason__name', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    list_filter = (
        'status', 'deleted', 'reason', 'created_on', 'approved_on', 'adjusted_on', 'rejected_on')
    readonly_fields = (
        'id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by', 'debit_note_number','broker')
    actions = ['export_csv']

    def booking_id(self, instance):
        return format_html('<br/>'.join([booking.booking_id for booking in instance.bookings.all()]))

    def delete_model(self, request, obj):
        obj.deleted = True
        obj.deleted_on = datetime.now()
        obj.changed_by = request.user
        super().delete_model(request, obj)

    def save_model(self, request, obj, form, change):
        obj.changed_by = request.user
        super().save_model(request, obj, form, change)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.deleted:
            return False
        return True

    def export_csv(modeladmin, request, queryset):
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=DNS {}.csv'.format(
            datetime.now().strftime('%d-%b-%Y %H:%M %s'))
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow(
            ['id', 'debit_note_number', 'booking_id', 'supplier', 'debit_amount', 'adjusted_amount', 'reason',
             'status', 'created_on', 'approved_on', 'adjusted_on', 'remarks'])
        for instance in queryset:
            writer.writerow([
                instance.id,
                instance.debit_note_number,
                '\n'.join([booking.booking_id for booking in instance.bookings.all()]),
                instance.accounting_supplier.name if isinstance(instance.accounting_supplier, Supplier) else None,
                instance.debit_amount,
                instance.adjusted_amount,
                instance.reason.name if isinstance(instance.reason, CreditDebitNoteReason) else None,
                instance.get_status_display(),
                instance.created_on.date() if instance.created_on else None,
                instance.approved_on.date() if instance.approved_on else None,
                instance.adjusted_on.date() if instance.adjusted_on else None,
                instance.remarks
            ])
        return response

    export_csv.short_description = u"Export CSV"


class ManualBookingSummaryAdmin(admin.ModelAdmin):
    list_display = ['id', 'user', 'vehicle', 'summary', 'updated_on']
    search_fields = ['id', 'user__username', 'vehicle__vehicle_number']
    readonly_fields = ('id', 'deleted_on', 'deleted', 'created_by', 'created_on', 'updated_on', 'changed_by')


class BookingStatusColorAdmin(admin.ModelAdmin):
    list_display = ('id', 'color_code', 'keyword')


class ManualBookingS3UploadAdmin(admin.ModelAdmin):
    list_display = ['id', 'booking_id', 'lr_number', 'is_valid', 'link']
    search_fields = ['id', 'booking__booking_id', 'booking__vehicle__vehicle_number', 'booking__lr_numbers__lr_number']
    list_filter = ['is_valid']
    readonly_fields = ['id', 'booking', 's3_upload', 'deleted_on', 'deleted', 'created_on', 'updated_on']

    def booking_id(self, obj):
        if isinstance(obj.booking, ManualBooking):
            return obj.booking.booking_id
        return ''

    def lr_number(self, obj):
        if isinstance(obj.booking, ManualBooking):
            return format_html('<br/>'.join([lr.lr_number for lr in obj.booking.lr_numbers.all()]))

    def link(self, obj):
        if isinstance(obj.s3_upload, S3Upload):
            return format_html('<a href="{}">download</a>'.format(obj.s3_upload.public_url()))
        else:
            return '-'

    link.short_description = 'Link'

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False


class LrS3UploadAdmin(admin.ModelAdmin):
    list_display = ['id', 'lr', 'booking_id', 'is_valid', 'link']
    search_fields = ['id', 'lr_number__lr_number', 'lr_number__booking__booking_id',
                     'lr_number__booking__vehicle__vehicle_number']
    list_filter = ('is_valid',)
    readonly_fields = ['id', 'lr_number', 's3_upload', 'deleted_on', 'deleted', 'created_on', 'updated_on']

    def booking_id(self, obj):
        if isinstance(obj.lr_number, LrNumber) and isinstance(obj.lr_number.booking, ManualBooking):
            return obj.lr_number.booking.booking_id
        return '-'

    def lr(self, obj):
        if isinstance(obj.lr_number, LrNumber):
            return obj.lr_number.lr_number
        return '-'

    def link(self, obj):
        if isinstance(obj.s3_upload, S3Upload):
            return format_html('<a href="{}">download</a>'.format(obj.s3_upload.public_url()))
        else:
            return '-'

    link.short_description = 'Link'

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False


class DataTablesFilterAdmin(admin.ModelAdmin):
    list_display = ['id', 'table_name', 'criteria_prettified']
    search_fields = ('id', 'table_name', 'criteria')
    readonly_fields = ['id', 'deleted_on', 'deleted', 'created_on', 'updated_on', 'created_by', 'changed_by']
    list_filter = ('table_name',)

    def has_delete_permission(self, request, obj=None):
        return False

    def criteria_prettified(self, instance):
        """Function to display pretty version of our data"""

        # Convert the data to sorted, indented JSON
        response = json.dumps(instance.criteria, sort_keys=True, indent=2)

        # Truncate the data. Alter as needed
        response = response

        # Get the Pygments formatter
        formatter = HtmlFormatter(style='colorful')

        # Highlight the data
        response = highlight(response, JsonLexer(), formatter)

        # Get the stylesheet
        style = "<style>" + formatter.get_style_defs() + "</style><br>"

        # Safe the output
        return mark_safe(style + response)

    criteria_prettified.short_description = 'data prettified'


admin.site.register(models.ManualBooking, ManualBookingAdmin)
admin.site.register(models.BookingStatusColor, BookingStatusColorAdmin)
admin.site.register(models.ManualBookingSummary, ManualBookingSummaryAdmin)
admin.site.register(models.Invoice, InvoiceAdmin)
# admin.site.register(models.ToPayInvoice, ToPayInvoiceAdmin)
admin.site.register(models.LrNumber, LrNumberAdmin)
admin.site.register(models.InWardPayment, InwardPaymentAdmin)
admin.site.register(models.OutWardPayment, OutWardPaymentAdmin)
admin.site.register(models.OutWardPaymentBill, OutWardPaymentBillAdmin)
admin.site.register(models.NotifyCompletedTaskEmail)
admin.site.register(models.DeletedData, DeletedDataAdmin)
admin.site.register(models.PendingInwardPaymentEntry, PendingInwardPaymentEntryAdmin)
admin.site.register(models.InvoiceSummary, InvoiceSummaryAdmin)
admin.site.register(models.RejectedPOD)
admin.site.register(models.CreditDebitNoteReason, CreditDebitNoteReasonAdmin)
admin.site.register(models.CreditNoteCustomer, CreditNoteCustomerAdmin)
admin.site.register(models.DebitNoteCustomer, DebitNoteCustomerAdmin)
admin.site.register(models.DebitNoteSupplier, DebitNoteSupplierAdmin)
admin.site.register(models.CreditNoteSupplier, CreditNoteSupplierAdmin)
admin.site.register(models.CreditNoteCustomerDirectAdvance, CreditNoteCustomerDirectAdvanceAdmin)
admin.site.register(models.ManualBookingS3Upload, ManualBookingS3UploadAdmin)
admin.site.register(models.LrS3Upload, LrS3UploadAdmin)
admin.site.register(models.DataTablesFilter, DataTablesFilterAdmin)
